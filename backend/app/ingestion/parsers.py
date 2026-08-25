"""Content parsers, keyed by document *kind*.

Supported kinds: html, pdf (text), pdf (scanned → OCR), xlsx/xls, json, rss,
xml, text. The kind is detected from an explicit adapter hint, then the MIME
type, then the URL extension, then by sniffing the bytes.

Parsing is derived output — a parse failure never discards the raw document
(the runner stores raw first). Heavy backends (pypdf, OCR) are imported lazily
and are injectable so the pipeline is testable without binaries.
"""
from __future__ import annotations

import csv
import io
import json
from collections.abc import Callable

from app.ingestion.errors import ParseError
from app.ingestion.types import FetchedDocument, ParsedContent

# Injectable backends.
PdfTextBackend = Callable[[bytes], str]
OcrEngine = Callable[[bytes], str]

_MIME_TO_KIND = {
    "text/html": "html",
    "application/xhtml+xml": "html",
    "application/pdf": "pdf",
    "application/json": "json",
    "application/rss+xml": "rss",
    "application/atom+xml": "rss",
    "text/xml": "xml",
    "application/xml": "xml",
    "text/csv": "csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xlsx",
    "text/plain": "text",
}

_EXT_TO_KIND = {
    ".html": "html", ".htm": "html",
    ".pdf": "pdf",
    ".json": "json",
    ".rss": "rss", ".atom": "rss",
    ".xml": "xml",
    ".csv": "csv",
    ".xlsx": "xlsx", ".xls": "xlsx",
    ".txt": "text",
}


def detect_kind(doc: FetchedDocument, hint: str | None = None) -> str:
    if hint:
        return hint
    mime = (doc.metadata.content_type or "").lower()
    if mime in _MIME_TO_KIND:
        return _MIME_TO_KIND[mime]

    url = doc.source_url.lower().split("?")[0]
    for ext, kind in _EXT_TO_KIND.items():
        if url.endswith(ext):
            return kind

    # Sniff bytes.
    head = doc.content[:16].lstrip()
    if head.startswith(b"%PDF"):
        return "pdf"
    if head.startswith(b"PK\x03\x04"):
        return "xlsx"  # OOXML is a zip
    if head[:1] in (b"{", b"["):
        return "json"
    if head.startswith(b"<?xml") or head.startswith(b"<rss") or head.startswith(b"<feed"):
        return "rss" if (b"<rss" in doc.content[:256] or b"<feed" in doc.content[:256]) else "xml"
    if head.startswith(b"<"):
        return "html"
    return "text"


# --------------------------------------------------------------------------- #
# Individual parsers
# --------------------------------------------------------------------------- #
def parse_html(doc: FetchedDocument) -> ParsedContent:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(doc.content, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    title = soup.title.string.strip() if soup.title and soup.title.string else doc.metadata.title
    text = "\n".join(line for line in (soup.get_text("\n").splitlines()) if line.strip())
    return ParsedContent(parser_name="html", text=text, title=title)


def parse_json(doc: FetchedDocument) -> ParsedContent:
    try:
        data = json.loads(doc.content.decode("utf-8"))
    except (ValueError, UnicodeDecodeError) as exc:
        raise ParseError(f"invalid JSON: {exc}") from exc
    return ParsedContent(parser_name="json", structured=data, title=doc.metadata.title)


def parse_rss(doc: FetchedDocument) -> ParsedContent:
    import feedparser

    feed = feedparser.parse(doc.content)
    entries = [
        {
            "title": e.get("title"),
            "link": e.get("link"),
            "summary": e.get("summary"),
            "published": e.get("published"),
            "id": e.get("id"),
        }
        for e in feed.get("entries", [])
    ]
    feed_title = feed.get("feed", {}).get("title") or doc.metadata.title
    text = "\n\n".join(f"{e['title']}\n{e['summary'] or ''}" for e in entries)
    return ParsedContent(parser_name="rss", structured=entries, text=text, title=feed_title)


def parse_xml(doc: FetchedDocument) -> ParsedContent:
    # Kept simple: expose raw text; structured XML parsing is source-specific.
    return ParsedContent(parser_name="xml", text=doc.content.decode("utf-8", errors="replace"))


def parse_csv(doc: FetchedDocument) -> ParsedContent:
    text = doc.content.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return ParsedContent(parser_name="csv", structured=rows, text=text)


def parse_excel(doc: FetchedDocument) -> ParsedContent:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(doc.content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of errors
        raise ParseError(f"invalid spreadsheet: {exc}") from exc
    sheets: dict[str, list[list]] = {}
    for ws in wb.worksheets:
        sheets[ws.title] = [list(row) for row in ws.iter_rows(values_only=True)]
    text_lines = []
    for name, rows in sheets.items():
        text_lines.append(f"# {name}")
        for row in rows:
            text_lines.append("\t".join("" if c is None else str(c) for c in row))
    return ParsedContent(parser_name="excel", structured=sheets, text="\n".join(text_lines))


def _default_pdf_text_backend(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def default_ocr_engine(data: bytes) -> str:  # pragma: no cover - requires binaries
    """Real OCR backend. Requires the ``pytesseract`` + ``pdf2image`` packages
    and the Tesseract + Poppler system binaries."""
    try:
        import pytesseract
        from pdf2image import convert_from_bytes
    except ImportError as exc:
        raise ParseError(
            "OCR requested but pytesseract/pdf2image not installed"
        ) from exc
    images = convert_from_bytes(data)
    return "\n".join(pytesseract.image_to_string(img) for img in images)


def parse_pdf(
    doc: FetchedDocument,
    *,
    pdf_text_backend: PdfTextBackend | None = None,
    ocr_engine: OcrEngine | None = None,
    ocr_min_chars: int = 20,
) -> ParsedContent:
    """Parse a PDF. If the text layer is (near-)empty the PDF is treated as
    *scanned* and OCR is used when an engine is available."""
    backend = pdf_text_backend or _default_pdf_text_backend
    try:
        text = backend(doc.content)
    except Exception as exc:
        raise ParseError(f"could not read PDF: {exc}") from exc

    scanned = len(text.strip()) < ocr_min_chars
    if scanned:
        if ocr_engine is not None:
            text = ocr_engine(doc.content)
            return ParsedContent(parser_name="pdf_ocr", text=text, extra={"scanned": True})
        # No OCR engine: preserve raw, flag for later OCR. Not a failure.
        return ParsedContent(
            parser_name="pdf", text=(text.strip() or None), extra={"scanned": True, "needs_ocr": True}
        )
    return ParsedContent(parser_name="pdf", text=text, extra={"scanned": False})


def parse_text(doc: FetchedDocument) -> ParsedContent:
    return ParsedContent(parser_name="text", text=doc.content.decode("utf-8", errors="replace"))


# --------------------------------------------------------------------------- #
# Dispatch
# --------------------------------------------------------------------------- #
def parse_document(
    doc: FetchedDocument,
    *,
    hint: str | None = None,
    pdf_text_backend: PdfTextBackend | None = None,
    ocr_engine: OcrEngine | None = None,
) -> ParsedContent:
    kind = detect_kind(doc, hint)
    if kind == "html":
        return parse_html(doc)
    if kind == "json":
        return parse_json(doc)
    if kind == "rss":
        return parse_rss(doc)
    if kind == "xml":
        return parse_xml(doc)
    if kind == "csv":
        return parse_csv(doc)
    if kind == "xlsx":
        return parse_excel(doc)
    if kind == "pdf":
        return parse_pdf(doc, pdf_text_backend=pdf_text_backend, ocr_engine=ocr_engine)
    return parse_text(doc)
